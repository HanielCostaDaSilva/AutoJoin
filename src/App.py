import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import model
import formatTable as ft

#= = = Variáveis 

# Obter o diretório atual do arquivo de script
diretorio_atual = os.path.dirname(os.path.realpath(__file__))
func_path = os.path.join(diretorio_atual,"..","data","func.xlsx")
func_atualizado_path = os.path.join(diretorio_atual,"..","data","func_atualizado.xlsx")
func_mesclado_path = os.path.join(diretorio_atual,"..","data","func_final.xlsx")

func_df = model.Dataframe(func_path)
func_atualizado_df = model.Dataframe(func_atualizado_path)

#Coletando a coluna PK do DF

columns_df = func_df.columns

colum_pk_index =-1

#pedimos a coluna que será a chave primária do DF
while colum_pk_index <= 0 or colum_pk_index > len(columns_df):
    print("digite o índice de uma coluna para ser a chave primária ")
    
    for i, column in enumerate(columns_df):
        print(f"{i + 1} = {column}" )  
    
    try:  
        colum_pk_index= int(input("=>"))
    except Exception as E:
        print(E)

#com o índice, coletamos a coluna primária
pk_column = columns_df[colum_pk_index - 1]

print(f" coluna {pk_column} escolhida")
input()

#Adicionando a coluna que será responsável para saber se vai alterar dataframe
situacao_col= "SITUACAO_DATAFRAME"
func_df.add_column(situacao_col)

#Adicionamos as colunas que estão na tabela original, mas não estão na tabela atualizada
for col in func_df.columns:

    func_atualizado_df.add_column(col)

#Posiveis situações
situacao_lista =["ALTERADO","ADICIONADO"]

#Percorremos o dataframe dos funcionarios atualizados
for _, func_atualizado in func_atualizado_df.get_iterrows():
        
    #Só realize a operação caso a situação seja PENDENTE
    if func_atualizado[situacao_col] in situacao_lista:
        continue    
    
    # Encontra os funcionários correspondentes no DataFrame original
    func_index_list = func_df.get_index_by_colum(pk_column, func_atualizado[pk_column]) 
    
    # Se o funcionário estiver presente no DataFrame original, UMA OU + VEZEZ atualizar todos os valores em branco
    if len(func_index_list) > 0:
        
        # Encontra as  duplicadas do funcionário no DataFrame atualizado
        func_atualiz_index_duplic = func_atualizado_df.get_index_by_colum(pk_column, func_atualizado[pk_column])
        duplic_index=0    
        
        for index in func_index_list:
            
            alterado=False
            
            for coluna in func_df.columns:
                # Se o valor da coluna no DataFrame original estiver em branco, preencher com o valor correspondente do DataFrame atualizado
            
                if func_df.get_column(index , coluna) == "":
                    func_df.alter_row(index, coluna, func_atualizado[coluna].upper())
                    alterado=True
             
            if alterado:  
                    func_df.alter_row(index, situacao_col, situacao_lista[0])
        
        #CASO ESPECIAL
        # O funcionário está presente no DataFrame ATUALIZADO, DUAS OU + VEZEZ; ADICIONE o valor no dataframe original

        if len(func_atualiz_index_duplic) > 1:
            
            #Adicione o valor no dataframe a partir da primeira duplicada;
            for func_dup_index in func_atualiz_index_duplic[1:]:
                func_dup = func_atualizado_df.alter_row(func_dup_index, situacao_col, situacao_lista[0])
                
                func_df.add_row(**func_dup)

                 
                
    #Caso não encontre o funcionario no dataframe, adicione-o no dataframe original
    else:
        func_atualizado.loc[situacao_col] = situacao_lista[1]
        func_df.add_row(**func_atualizado)
    
    #print(func_atualizado)
        
# Salvar o DataFrame mesclado
try:
    func_df.save(func_mesclado_path)
except PermissionError as PE:
    print("Não foi acessar a tabela...\n Tente fechar o arquivo, caso esteja aberto")
    print(PE)
    input()
    # Clearing the Screen
    os.system('cls')
   
# Carregar o arquivo Excel mesclado
workbook = load_workbook(func_mesclado_path)
ws = workbook.active

#quantidade máxima de linhas
quant_func = len(func_df)
#quantidade máxima de colunas
quant_col = len(func_df.columns)

""" #colunas não obrigatorias
colum_no_obligate = [
func_df.columns.get_loc("NUMERO"),
func_df.columns.get_loc("COMPLEMENTO"),
func_df.columns.get_loc("FONE")
] """

#Amarelo se as tabelas estão preenchidas corretamente
yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

#Vermelha se estiver faltando alguma informação OBRIGATORIA na coluna
#red = PatternFill(start_color="DE4740", end_color="DE4740", fill_type="solid")

#VERDE se for um funcionario for adicionado a planilha original
green = PatternFill(start_color="40DE47", end_color="40DE47", fill_type="solid")

# Iterar sobre as células alteradas e aplicar o estilo de preenchimento amarelo nelas
for row in ws.iter_rows(min_row=2, max_row= quant_func + 1, min_col=1, max_col=quant_col):
    #Checa se a linha foi alterada
    
    if row[func_df.columns.get_loc(situacao_col)].value == situacao_lista[0]:
        #pinta as linhas alteradas de amarelo
        ft.paintRow(row,yellow)   
        
    elif row[func_df.columns.get_loc(situacao_col)].value == situacao_lista[1]:
        ft.paintRow(row,green)
        
# Salvar as alterações no arquivo Excel mesclado
workbook.save(func_mesclado_path) 

print("*-"*30)
print("Programa finalizado (pressione Enter)....")
print("*-"*30)
input()